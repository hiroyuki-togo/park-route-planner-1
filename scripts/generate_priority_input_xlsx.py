"""優先度（default_priority）の見直し用 Excel 雛形を生成する。

各アトラクションの現在の default_priority を B 列に pre-fill しておくので、
東郷さんは変えたいものだけ上書き入力 → 共有すれば
`scripts/import_priority_from_xlsx.py` で取り込める。

出力: data/priority_input.xlsx
"""
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="4472C4")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=11)
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")  # 入力欄を薄黄色
INFO_FONT = Font(name="Arial", size=9, color="595959")
ID_FONT = Font(name="Arial", size=9, color="808080")

AREA_ORDER = [
    "ワールドバザール", "アドベンチャーランド", "ウエスタンランド",
    "クリッターカントリー", "ファンタジーランド", "トゥーンタウン", "トゥモローランド",
]
PASS_LABEL = {"dpa": "DPA", "priority": "プライオリティ"}


def build_info(a):
    """補足情報を短くまとめた文字列を作る（tier / pass / 待ち / 屋内外）。"""
    parts = [f"tier {a.get('popularity_tier','?')}"]
    pt = a.get("pass_type")
    if pt:
        parts.append(PASS_LABEL.get(pt, pt))
    parts.append(f"avg待ち {a.get('avg_wait_min','?')}分")
    parts.append("屋外" if a.get("outdoor") else "屋内")
    return " / ".join(parts)


def build_sheet(wb, attractions):
    ws = wb.create_sheet("優先度")
    headers = ["アトラクション名", "優先度（0〜5）", "補足情報", "エリア", "ID（編集不要）"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for row, a in enumerate(attractions, start=2):
        ws.cell(row=row, column=1, value=a["name"]).font = BODY_FONT
        cb = ws.cell(row=row, column=2, value=a.get("default_priority"))
        cb.fill = INPUT_FILL
        cb.font = BODY_FONT
        cb.alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=3, value=build_info(a)).font = INFO_FONT
        ws.cell(row=row, column=4, value=a["area"]).font = BODY_FONT
        cid = ws.cell(row=row, column=5, value=a["id"])
        cid.font = ID_FONT

    widths = [38, 14, 44, 22, 22]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    return ws


def build_instructions(wb):
    ws = wb.create_sheet("使い方", 0)
    lines = [
        ("東京ディズニーランド 優先度設定シート", True),
        ("", False),
        ("【目的】", True),
        ("各アトラクションの「優先度（default_priority）」をご家族の好みに合わせて事前設定してください。", False),
        ("ここで設定した値が、当日の UI で初期値として使われます（UI で個別に変更も可能）。", False),
        ("", False),
        ("【手順】（変えたいものだけ書き換える方式）", True),
        ("1. 「優先度」シートを開く（34 行ある）", False),
        ("2. B 列に現在の優先度が pre-fill されている。変えたい行だけ B 列を上書き", False),
        ("3. 全部変える必要はない。そのままで良い行は触らなくて OK", False),
        ("4. 入力後、Claude に「入力完了」と伝える → 自動で data/attractions.json に反映", False),
        ("", False),
        ("【優先度の意味】", True),
        ("5 = 必ず乗りたい（must-visit 候補）", False),
        ("4 = できれば乗りたい", False),
        ("3 = 余裕あれば乗る（デフォルト的位置付け）", False),
        ("2 = あまり優先しない", False),
        ("1 = 候補に出るだけ（ほぼ立ち寄り扱い）", False),
        ("0 = 候補から除外（UI で「必ず乗る」と矛盾しないようガード済）", False),
        ("", False),
        ("【補足情報（C 列）の見方】", True),
        ("tier S/A/B/C = 人気度（S が最高）。スコア計算で人気度が高いほど価値が大きい", False),
        ("DPA = ディズニー・プレミアアクセス対象 / プライオリティ = プライオリティパス対象", False),
        ("avg待ち = 平均待ち時間の推測値（実値は当日 Queue-Times で取得）", False),
        ("屋外 / 屋内 = 雨天モード時の待ち時間補正に影響", False),
        ("", False),
        ("【注意事項】", True),
        ("・A 列（名前）と E 列（ID）は編集しないでください。マッピングの鍵になります", False),
        ("・B 列は 0〜5 の整数のみ。それ以外を入れると取り込みエラーになります", False),
        ("・「必ず乗る」に既に該当しているもの（pass_type=DPA / プライオリティ）は当日 UI で優先される仕組みなので、ここの数値は補助的", False),
    ]
    for i, (text, is_header) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if is_header:
            c.font = Font(name="Arial", bold=True, size=12, color="1F4E78")
        else:
            c.font = Font(name="Arial", size=11)
    ws.column_dimensions["A"].width = 110
    return ws


def main():
    raw = json.loads(Path("data/attractions.json").read_text())
    attractions = raw["attractions"]

    def sort_key(a):
        area_idx = AREA_ORDER.index(a["area"]) if a["area"] in AREA_ORDER else 99
        # エリア順 → エリア内は現在 priority 降順 → 名前
        return (area_idx, -a.get("default_priority", 0), a["name"])

    attractions.sort(key=sort_key)

    wb = Workbook()
    wb.remove(wb.active)
    build_instructions(wb)
    build_sheet(wb, attractions)

    out = Path("data/priority_input.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote xlsx: {out} (attractions={len(attractions)})")


if __name__ == "__main__":
    main()
