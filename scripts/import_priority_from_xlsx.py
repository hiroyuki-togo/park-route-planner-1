"""東郷さんから戻ってきた Excel（priority_input.xlsx）を読み込み、
default_priority を data/attractions.json に反映する。

使い方:
    python scripts/import_priority_from_xlsx.py [xlsx_path]

デフォルトは data/priority_input.xlsx を読み込む。
ID 列（E 列）で JSON エントリと突き合わせるため、A 列の名前が変わっていても安全。
"""
import json
import sys
from pathlib import Path

from openpyxl import load_workbook


VALID_RANGE = (0, 5)
SHEET_NAME = "優先度"


def parse_priority(raw):
    """B 列の値を 0〜5 の整数にする。空欄は None（変更なし扱い）。"""
    if raw is None:
        return None
    if isinstance(raw, str) and not raw.strip():
        return None
    try:
        n = int(raw)
    except (ValueError, TypeError):
        raise ValueError(f"優先度が整数ではない: {raw!r}（期待: 0〜5）")
    if not (VALID_RANGE[0] <= n <= VALID_RANGE[1]):
        raise ValueError(f"優先度が範囲外: {n}（期待: 0〜5）")
    return n


def read_sheet(ws):
    """シートから {id: priority or None} を返す + フォーマットエラー一覧。"""
    result = {}
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        name, priority_raw, _info, _area, id_ = row[0], row[1], row[2], row[3], row[4]
        if not id_:
            continue
        try:
            result[id_] = parse_priority(priority_raw)
        except ValueError as e:
            errors.append(f"  - {id_} ({name}): {e}")
    return result, errors


def apply_to_json(json_path, priorities):
    """JSON の attractions 配列に default_priority を流し込む。"""
    raw = json.loads(json_path.read_text())
    updated, unchanged, missing_id = 0, 0, []
    for entry in raw["attractions"]:
        new_pri = priorities.get(entry["id"])
        if new_pri is None:
            continue
        old_pri = entry.get("default_priority")
        if new_pri == old_pri:
            unchanged += 1
            continue
        entry["default_priority"] = new_pri
        updated += 1
        print(f"  ✏️  {entry['id']:22} {entry['name'][:30]:30} {old_pri} → {new_pri}")
    json_ids = {e["id"] for e in raw["attractions"]}
    for xid in priorities:
        if xid not in json_ids:
            missing_id.append(xid)
    json_path.write_text(json.dumps(raw, ensure_ascii=False, indent=2) + "\n")
    return updated, unchanged, missing_id


def main(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        sys.exit(f"❌ シート {SHEET_NAME!r} が見つかりません")

    priorities, errors = read_sheet(wb[SHEET_NAME])
    if errors:
        print("❌ フォーマットエラー（その項目は反映スキップ）:")
        for e in errors:
            print(e)
        print()

    json_path = Path("data/attractions.json")
    print("=== 変更検出 ===")
    updated, unchanged, missing = apply_to_json(
        json_path,
        {k: v for k, v in priorities.items() if v is not None},
    )
    print()
    print("=== 反映結果 ===")
    print(f"更新: {updated} 件 / 変更なし: {unchanged} 件")
    if missing:
        print(f"⚠️  Excel にあって JSON にない ID: {missing}")
    print()
    print("確認: .venv/bin/pytest -q")


if __name__ == "__main__":
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/priority_input.xlsx")
    if not xlsx.exists():
        sys.exit(f"ファイルが見つかりません: {xlsx}")
    main(xlsx)
