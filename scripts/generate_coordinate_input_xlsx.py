"""座標入力用の Excel 雛形を生成する。
東郷さんが Google マップから取得した座標を B 列にペーストして共有してもらう。

`--missing-only` フラグを付けると、lat または lng が null の行だけを出力し、
別ファイル `data/coordinate_input_missing.xlsx` に保存する。マスタへの追加分
だけを再入力してもらうときに使う（既存の入力済 xlsx を上書きしないため）。
"""
import argparse
import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


HEADER_FILL = PatternFill("solid", start_color="4472C4")
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF", size=11)
BODY_FONT = Font(name="Arial", size=11)
INPUT_FILL = PatternFill("solid", start_color="FFF2CC")  # 入力欄を薄黄色で目立たせる
ID_FONT = Font(name="Arial", size=9, color="808080")


def build_sheet(wb, sheet_name, header_label, items):
    """items: list of (id, name, area)"""
    ws = wb.create_sheet(sheet_name)
    headers = [header_label, "座標（緯度, 経度）", "エリア", "ID（編集不要）"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for row, (id_, name, area) in enumerate(items, start=2):
        ws.cell(row=row, column=1, value=name).font = BODY_FONT
        cb = ws.cell(row=row, column=2, value=None)
        cb.fill = INPUT_FILL
        cb.font = BODY_FONT
        ws.cell(row=row, column=3, value=area).font = BODY_FONT
        cid = ws.cell(row=row, column=4, value=id_)
        cid.font = ID_FONT

    widths = [40, 30, 22, 24]
    for col, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A2"
    return ws


def build_instructions(wb):
    ws = wb.create_sheet("使い方", 0)
    lines = [
        ("東京ディズニーランド 座標入力シート", True),
        ("", False),
        ("【目的】", True),
        ("各アトラクション・レストランの入口座標を Google マップから取得し、B 列に入力してください。", False),
        ("入力済みの Excel を Claude に共有すると、自動で JSON ファイル（data/attractions.json / data/restaurants.json）に反映します。", False),
        ("", False),
        ("【手順】（1 件あたり 30 秒）", True),
        ("1. Google マップ（https://www.google.com/maps）を開く", False),
        ("2. 検索窓に「東京ディズニーランド ○○」と入力（例：東京ディズニーランド ホーンテッドマンション）", False),
        ("3. 地図上のピンを右クリック → コンテキストメニュー最上部に「35.6332, 139.8801」のような座標が表示される", False),
        ("4. 座標をクリックでコピー → Excel の B 列のセルにペースト", False),
        ("5. シート下部のタブで「アトラクション」と「レストラン」を切り替えながら全件入力", False),
        ("", False),
        ("【入力フォーマット】", True),
        ("B 列にはカンマ区切りでそのまま貼り付けて OK。例：35.6332456, 139.8801234", False),
        ("小数点は 4 桁あれば十分（1m 弱の精度）。", False),
        ("Google マップが返す形式そのままで OK（Claude 側でパースします）。", False),
        ("", False),
        ("【注意事項】", True),
        ("・A 列（名前）と D 列（ID）は編集しないでください。Claude が JSON にマッピングする際の鍵になります。", False),
        ("・入口が複数あるアトラクションは、キューが伸びる側の入口を選ぶのがベター（ただし ±20m はルート計算上影響なし）。", False),
        ("・どうしても見つからない場合は B 列を空欄のままにして OK。Claude 側で再確認します。", False),
        ("・美女と野獣（プライオリティパス対応）の現状運用が変わっていたら、メモ欄や口頭で教えてください。", False),
    ]
    for i, (text, is_header) in enumerate(lines, start=1):
        c = ws.cell(row=i, column=1, value=text)
        if is_header:
            c.font = Font(name="Arial", bold=True, size=12, color="1F4E78")
        else:
            c.font = Font(name="Arial", size=11)
    ws.column_dimensions["A"].width = 110
    return ws


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--missing-only",
        action="store_true",
        help="lat または lng が null の行のみ出力（出力先は coordinate_input_missing.xlsx）",
    )
    args = parser.parse_args()

    attractions_raw = json.loads(Path("data/attractions.json").read_text())
    restaurants_raw = json.loads(Path("data/restaurants.json").read_text())

    def _is_missing(item):
        return item.get("lat") is None or item.get("lng") is None

    if args.missing_only:
        attractions = [
            (a["id"], a["name"], a["area"])
            for a in attractions_raw["attractions"]
            if _is_missing(a)
        ]
        restaurants = [
            (r["id"], r["name"], r["area"])
            for r in restaurants_raw["restaurants"]
            if _is_missing(r)
        ]
    else:
        attractions = [(a["id"], a["name"], a["area"]) for a in attractions_raw["attractions"]]
        restaurants = [(r["id"], r["name"], r["area"]) for r in restaurants_raw["restaurants"]]

    # エリア順に並べ替えると検索効率が上がる
    area_order = [
        "ワールドバザール", "アドベンチャーランド", "ウエスタンランド",
        "クリッターカントリー", "ファンタジーランド", "トゥーンタウン", "トゥモローランド",
    ]
    def area_key(item):
        try:
            return area_order.index(item[2])
        except ValueError:
            return len(area_order)
    attractions.sort(key=area_key)
    restaurants.sort(key=area_key)

    wb = Workbook()
    wb.remove(wb.active)  # default sheet を削除
    build_instructions(wb)
    build_sheet(wb, "アトラクション", "アトラクション名", attractions)
    build_sheet(wb, "レストラン", "レストラン名", restaurants)

    out_name = "coordinate_input_missing.xlsx" if args.missing_only else "coordinate_input.xlsx"
    out = Path("data") / out_name
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out)
    print(f"Wrote xlsx: {out} (attractions={len(attractions)}, restaurants={len(restaurants)})")


if __name__ == "__main__":
    main()
