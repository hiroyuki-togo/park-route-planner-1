"""東郷さんから戻ってきた Excel を読み込み、座標を JSON に反映する。

使い方:
    python scripts/import_coordinates_from_xlsx.py [xlsx_path]

デフォルトは data/coordinate_input.xlsx を読み込む。
ID 列（D 列）で JSON エントリと突き合わせるため、A 列の名前が変わっていても安全。
"""
import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


COORD_RE = re.compile(r"^\s*(-?\d+(?:\.\d+)?)\s*,\s*(-?\d+(?:\.\d+)?)\s*$")
# TDL は北緯 35.62-35.64 / 東経 139.87-139.89 の範囲。緩めに見る
TDL_LAT_RANGE = (35.60, 35.66)
TDL_LNG_RANGE = (139.85, 139.91)


def parse_coord(raw):
    """Google マップ形式 '35.6332, 139.8801' を (lat, lng) にする。空欄は None。"""
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    m = COORD_RE.match(s)
    if not m:
        raise ValueError(f"座標フォーマット不正: {raw!r}（期待: '35.6332, 139.8801'）")
    lat, lng = float(m.group(1)), float(m.group(2))
    if not (TDL_LAT_RANGE[0] <= lat <= TDL_LAT_RANGE[1]):
        raise ValueError(f"緯度が TDL 範囲外: {lat}（{raw!r}）")
    if not (TDL_LNG_RANGE[0] <= lng <= TDL_LNG_RANGE[1]):
        raise ValueError(f"経度が TDL 範囲外: {lng}（{raw!r}）")
    return lat, lng


def read_sheet(ws):
    """シートから {id: (lat, lng) or None} を返す。"""
    result = {}
    errors = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        name, coord_raw, _area, id_ = row[0], row[1], row[2], row[3]
        if not id_:
            continue
        try:
            result[id_] = parse_coord(coord_raw)
        except ValueError as e:
            errors.append(f"  - {id_} ({name}): {e}")
    return result, errors


def apply_to_json(json_path, key, coords):
    """JSON の attractions / restaurants 配列に座標を流し込む。"""
    raw = json.loads(json_path.read_text())
    updated, skipped, missing_id = 0, 0, []
    for entry in raw[key]:
        c = coords.get(entry["id"])
        if c is None:
            skipped += 1
            continue
        entry["lat"], entry["lng"] = c
        updated += 1
    # Excel 側にあって JSON 側にない ID
    json_ids = {e["id"] for e in raw[key]}
    for xid in coords:
        if xid not in json_ids:
            missing_id.append(xid)
    json_path.write_text(json.dumps(raw, ensure_ascii=False, indent=2) + "\n")
    return updated, skipped, missing_id


def main(xlsx_path):
    wb = load_workbook(xlsx_path, data_only=True)
    summary = []
    all_errors = []

    plan = [
        ("アトラクション", Path("data/attractions.json"), "attractions"),
        ("レストラン", Path("data/restaurants.json"), "restaurants"),
    ]
    for sheet_name, json_path, json_key in plan:
        if sheet_name not in wb.sheetnames:
            print(f"⚠️  シート {sheet_name!r} が見つかりません。スキップ", file=sys.stderr)
            continue
        coords, errors = read_sheet(wb[sheet_name])
        all_errors.extend([f"[{sheet_name}] {e}" for e in errors])
        # None（空欄）は反映しない
        non_null = {k: v for k, v in coords.items() if v is not None}
        updated, skipped, missing = apply_to_json(json_path, json_key, non_null)
        summary.append((sheet_name, updated, skipped, missing))

    if all_errors:
        print("❌ フォーマットエラーがありました（その項目はスキップ）:")
        for e in all_errors:
            print(e)
        print()

    print("=== 反映結果 ===")
    for sheet_name, updated, skipped, missing in summary:
        print(f"{sheet_name}: 更新 {updated} 件 / 空欄スキップ {skipped} 件")
        if missing:
            print(f"  ⚠️  Excel にあって JSON にない ID: {missing}")
    print()
    print("確認: .venv/bin/pytest tests/test_masters.py -v")


if __name__ == "__main__":
    xlsx = Path(sys.argv[1]) if len(sys.argv) > 1 else Path("data/coordinate_input.xlsx")
    if not xlsx.exists():
        sys.exit(f"ファイルが見つかりません: {xlsx}")
    main(xlsx)
